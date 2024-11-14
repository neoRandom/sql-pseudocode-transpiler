from models import DatabaseSchema
import os
import pandas as pd
import openpyxl as opx
import openpyxl.styles as opxs


def parse_json_to_obj(json_path: str) -> list[list[str]]:
    json_file = open(json_path, 'r', encoding="UTF-8")
    database: DatabaseSchema = DatabaseSchema.model_validate_json(json_file.read())

    obj_rows: list[list[str]] = []

    for table in database.table_list:
        obj_rows.append(["Tabela", table.name])
        obj_rows.append(["Descrição", table.description])
        obj_rows.append(["Observações", table.notes])
        obj_rows.append(["Campos"])
        obj_rows.append(["Nome", "Descrição", "Tipo de dado", "Tamanho", "Restrições"])

        for attribute in table.attribute_list:
            attr_name = attribute.name
            attr_desc = attribute.description
            attr_type = attribute.type
            attr_size = attribute.size
            modifiers = ", ".join(
                [
                    "FK" if m.startswith("fk")
                    else "PK, Identity" if m == "pk"
                    else m.capitalize()
                    for m in attribute.modifiers 
                ]
            ) 
            
            # Adiciona os dados da tabela em uma lista
            obj_rows.append([attr_name, attr_desc, attr_type, attr_size, modifiers])
        obj_rows.append([""])

    return obj_rows


def generate_excel(tables: list[list[str]], output_file: str) -> bool:
    # Creating the DataFrame
    df = pd.DataFrame(tables)
    
    # Creating the Excel file
    df.to_excel(output_file, index=False, sheet_name="main") # type: ignore
    
    # Opening the Excel file to do some alterations
    wb = opx.load_workbook(output_file)
    ws = wb["main"]

    # Defining the default font and the default border
    default_font = opxs.Font(name="Calibri (Body)", sz=12)
    default_border = opxs.Border(
        left=opxs.Side(style="thin"),
        right=opxs.Side(style="thin"),
        top=opxs.Side(style="thin"),
        bottom=opxs.Side(style="thin")
    )
    for row in range(2, ws.max_row):
        for cell in ws[row]:
            cell.font = default_font
            cell.border = default_border


    # Removing the index at row 1
    for j in range(5):
        ws[1][j].value = ""
        ws[1][j].border = opxs.Border()

    # Updating the other styles
    row_index = 2
    while row_index < ws.max_row:
        if ws[row_index][0].value != "Tabela":
            row_index += 1
            continue

        campos_row = row_index + 3
        attributes_row = row_index + 4

        font = opxs.Font(name="Calibri (Body)", sz=12, b=True)
        alignment = opxs.Alignment(horizontal="center", vertical="center")
        blue_fill = opxs.PatternFill(start_color="bdd7ee", end_color="bdd7ee", fill_type="solid")
        green_fill = opxs.PatternFill(start_color="c6e0b4", end_color="c6e0b4", fill_type="solid")

        # Setting the upper row (if it has one)
        if row_index > 5:
            upper_row = row_index - 1
            ws.merge_cells(start_row=upper_row, start_column=1, end_row=upper_row, end_column=5)
            ws.row_dimensions[upper_row].height = 25
            for cell in range(5):
                ws[upper_row][cell].border = opxs.Border()
        
        # Setting the first 3 rows
        for i in range(row_index, row_index + 3):
            ws.row_dimensions[i].height = 20
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
            ws[i][0].font = font
            ws[i][0].fill = blue_fill
            ws[i][0].alignment = alignment
            ws[i][1].alignment = opxs.Alignment(horizontal="left", vertical="center", indent=1)
        
        # Setting the 4th row
        ws.merge_cells(start_row=campos_row, start_column=1, end_row=campos_row, end_column=5)
        ws.row_dimensions[campos_row].height = 18
        ws[campos_row][0].font = font
        ws[campos_row][0].fill = green_fill
        ws[campos_row][0].alignment = alignment

        # Setting the 5th row
        ws.row_dimensions[attributes_row].height = 20
        for j in range(5):
            ws[attributes_row][j].font = font
            ws[attributes_row][j].fill = blue_fill
            ws[attributes_row][j].alignment = alignment
        
        row_index += 6
    
    # Calculate and set optimal column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                # Update max_length if the cell value length is greater
                max_length = max(max_length, len(str(cell.value)))
        # Adjust column width (approximate width in Excel is max_length * 1.2)
        adjusted_width = (max_length + 2) * 1.2  # Add padding and adjust for font
        ws.column_dimensions[column].width = adjusted_width
    
    # Saving alterations
    wb.save(output_file)

    return True


def transpile(json_path: str, *, update_files: bool, verbose_output: bool):
    output_file = json_path[:json_path.rfind(".")] + ".xlsx"

    if os.path.exists(output_file):
        if update_files:
            os.remove(output_file)
        else:
            return False
    
    if verbose_output:
        print("Generating the Data Dictionary (Excel)...")

    tables = parse_json_to_obj(json_path)
    if generate_excel(tables, output_file):
        if verbose_output:
            print("Data Dictionary (Excel) generated successfully.")
    else:
        if verbose_output:
            print("Data Dictionary not created/updated.")

